# =====================================================
# Template-UnitTestScript - Copyright (c) 2024
# Author: B. N. Andrei
# GitHub: Andrei6700
# Creation date: 21-10-2024
# Last update: 22-10-2024
# Version: 1.0
# 
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"),
# to deal in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies
# of the Software, and to permit persons to whom the Software is furnished to do so,
# subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all copies
# or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED,
# INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE
# AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM,
# DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE,
# ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
# =====================================================


# How to run: 
# 1. cmd from script location
# 2. py Template-UnitTestScript.py


from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font

# path ul tau:
file_path = r"your path"
workbook = load_workbook(file_path)

# alt tip de culoare nu merge, doar rgb
color_line_1_2 = PatternFill(start_color="91CCDB", end_color="91CCDB", fill_type="solid")  # RGB(145, 204, 219)
color_call_function = PatternFill(start_color="C3D696", end_color="C3D696", fill_type="solid")  # RGB(195, 214, 154)

# marginile
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# culoarea default
font_style = Font(name="Calibri", size=11, bold=False, color='000000')

# for pt a trece prin toate sheet-urile 
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    # linia 1 si 2
    for row in [1, 2]:
        for cell in sheet[row]:
            cell.fill = color_line_1_2  # adaugarea culorii
            cell.border = thin_border   # adaugarea marginii
            cell.font = font_style      # font

    # linia "CALL FUNCTION"
    for row in range(1, sheet.max_row + 1):
        for cell in sheet[row]:
            if cell.value == "CALL FUNCTION":
                for c in sheet[row]:  # adaugarea culorii, marinilor si a fontului
                    c.fill = color_call_function  # adaugarea culorii
                    c.border = thin_border        # adaugarea margini
                    c.font = font_style           # font
                break  

    # fontul/marginile pe toate celulelel, care nu sunt goale
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.value is not None: 
                cell.border = thin_border 
                cell.font = font_style  

    # ajustarea coloanelor in functie de dimensiunea textului
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        sheet.column_dimensions[column_letter].width = max_length + 2  # setarea latimii

# fisierul excel actualizat
# locatia unde sa se salveze fisieru' de output + numele
excel_output = r"path\name.xlsx"
workbook.save(excel_output)
print("fisierul tau a fost actualizat !")